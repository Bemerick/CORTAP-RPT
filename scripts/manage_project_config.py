#!/usr/bin/env python3
"""
Project Setup Configuration Management

This script manages project setup configuration data:
1. Exports JSON configuration to Excel template for easy editing
2. Imports Excel template back to JSON configuration
3. Validates configuration against schema

Usage:
    # Export JSON to Excel template
    python scripts/manage_project_config.py export --output config/project-setup-template.xlsx

    # Import Excel to JSON
    python scripts/manage_project_config.py import --input config/project-setup-template.xlsx --output config/project-setup.json

    # Validate JSON configuration
    python scripts/manage_project_config.py validate --config config/project-setup.json
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Missing required package: {e}")
    print("Install with: pip install pandas openpyxl")
    sys.exit(1)


class ProjectConfigManager:
    """Manage project setup configuration data"""

    def __init__(self, schema_path: str = "config/project-setup-schema.json"):
        self.schema_path = Path(schema_path)

    def export_to_excel(self, json_path: str, excel_path: str):
        """
        Export JSON configuration to Excel template with multiple sheets.

        Sheets:
        - Projects: Basic project info
        - Contractors: Contractor details
        - FTA_PM: FTA Program Manager details
        - Regional_Officer: Regional officer details
        - Attendees: Site visit attendees
        - Subrecipients: Subrecipient agencies
        """
        print(f"📤 Exporting {json_path} to Excel template...")

        # Load JSON
        with open(json_path, 'r') as f:
            config = json.load(f)

        projects = config.get('projects', [])

        if not projects:
            print("⚠️  No projects found in configuration")
            return

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Project Info
        self._create_project_info_sheet(wb, projects)

        # Sheet 2: Contractors
        self._create_contractors_sheet(wb, projects)

        # Sheet 3: FTA Program Managers
        self._create_fta_pm_sheet(wb, projects)

        # Sheet 4: Regional Officers
        self._create_regional_officer_sheet(wb, projects)

        # Sheet 5: Attendees
        self._create_attendees_sheet(wb, projects)

        # Sheet 6: Subrecipients
        self._create_subrecipients_sheet(wb, projects)

        # Sheet 7: Instructions
        self._create_instructions_sheet(wb)

        # Save workbook
        wb.save(excel_path)
        print(f"✅ Excel template saved to: {excel_path}")
        print(f"   Sheets: {', '.join(wb.sheetnames)}")

    def _create_project_info_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create Project Info sheet"""
        ws = wb.create_sheet("Project_Info")

        # Headers
        headers = [
            "project_id", "region_number", "review_type", "recipient_city_state",
            "recipient_acronym", "recipient_website", "fiscal_year",
            "site_visit_start_date", "site_visit_end_date", "site_visit_dates",
            "exit_conference_format", "report_date", "notes"
        ]

        ws.append(headers)
        self._style_header_row(ws, 1)

        # Data rows
        for project in projects:
            info = project.get('project_info', {})
            row = [
                project.get('project_id'),
                info.get('region_number'),
                info.get('review_type'),
                info.get('recipient_city_state'),
                info.get('recipient_acronym', ''),
                info.get('recipient_website', ''),
                info.get('fiscal_year'),
                info.get('site_visit_start_date'),
                info.get('site_visit_end_date'),
                info.get('site_visit_dates'),
                info.get('exit_conference_format'),
                info.get('report_date', ''),
                project.get('notes', '')
            ]
            ws.append(row)

        self._auto_size_columns(ws)

    def _create_contractors_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create Contractors sheet"""
        ws = wb.create_sheet("Contractors")

        headers = ["project_id", "lead_reviewer_name", "company_name",
                   "lead_reviewer_phone", "lead_reviewer_email"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for project in projects:
            contractor = project.get('contractor', {})
            row = [
                project.get('project_id'),
                contractor.get('lead_reviewer_name'),
                contractor.get('company_name'),
                contractor.get('lead_reviewer_phone'),
                contractor.get('lead_reviewer_email')
            ]
            ws.append(row)

        self._auto_size_columns(ws)

    def _create_fta_pm_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create FTA Program Manager sheet"""
        ws = wb.create_sheet("FTA_PM")

        headers = ["project_id", "name", "title", "phone", "email"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for project in projects:
            pm = project.get('fta_program_manager', {})
            row = [
                project.get('project_id'),
                pm.get('name'),
                pm.get('title'),
                pm.get('phone'),
                pm.get('email')
            ]
            ws.append(row)

        self._auto_size_columns(ws)

    def _create_regional_officer_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create Regional Officer sheet"""
        ws = wb.create_sheet("Regional_Officer")

        headers = ["project_id", "name", "title", "phone", "email"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for project in projects:
            ro = project.get('regional_officer', {})
            row = [
                project.get('project_id'),
                ro.get('name'),
                ro.get('title'),
                ro.get('phone'),
                ro.get('email')
            ]
            ws.append(row)

        self._auto_size_columns(ws)

    def _create_attendees_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create Attendees sheet"""
        ws = wb.create_sheet("Attendees")

        headers = ["project_id", "name", "organization", "title", "email"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for project in projects:
            project_id = project.get('project_id')
            attendees = project.get('attendees', [])

            if not attendees:
                # Add empty row for project
                ws.append([project_id, '', '', '', ''])
            else:
                for attendee in attendees:
                    row = [
                        project_id,
                        attendee.get('name'),
                        attendee.get('organization'),
                        attendee.get('title', ''),
                        attendee.get('email', '')
                    ]
                    ws.append(row)

        self._auto_size_columns(ws)

    def _create_subrecipients_sheet(self, wb: Workbook, projects: List[Dict]):
        """Create Subrecipients sheet"""
        ws = wb.create_sheet("Subrecipients")

        headers = ["project_id", "program_type", "name", "location"]
        ws.append(headers)
        self._style_header_row(ws, 1)

        for project in projects:
            project_id = project.get('project_id')
            subrecipients = project.get('subrecipients', {})

            for program_type in ['5307', '5310', '5311']:
                subs = subrecipients.get(program_type, [])

                if not subs:
                    # Add empty row
                    ws.append([project_id, program_type, '', ''])
                else:
                    for sub in subs:
                        row = [
                            project_id,
                            program_type,
                            sub.get('name'),
                            sub.get('location', '')
                        ]
                        ws.append(row)

        self._auto_size_columns(ws)

    def _create_instructions_sheet(self, wb: Workbook):
        """Create Instructions sheet"""
        ws = wb.create_sheet("Instructions", 0)  # Insert as first sheet

        instructions = [
            ["CORTAP Project Setup Configuration - Excel Template"],
            [""],
            ["INSTRUCTIONS:"],
            ["1. Edit the data in the sheets to configure your projects"],
            ["2. Do NOT modify the header row in any sheet"],
            ["3. Do NOT rename the sheets"],
            ["4. Use consistent project_id across all sheets"],
            ["5. After editing, save the file and import back to JSON using:"],
            ["   python scripts/manage_project_config.py import --input <file>.xlsx --output config/project-setup.json"],
            [""],
            ["SHEET DESCRIPTIONS:"],
            ["• Project_Info: Basic project and review information"],
            ["• Contractors: Lead reviewer and contractor company details"],
            ["• FTA_PM: FTA Program Manager assigned to the project"],
            ["• Regional_Officer: Regional Administrator information"],
            ["• Attendees: Site visit attendees (can have multiple rows per project)"],
            ["• Subrecipients: Subrecipient agencies by program type (5307, 5310, 5311)"],
            [""],
            ["FIELD FORMATS:"],
            ["• region_number: Integer 1-10"],
            ["• review_type: 'Triennial Review', 'State Management Review', or 'Combined Triennial and State Management Review'"],
            ["• recipient_city_state: Format: 'City, ST' (e.g., 'Seattle, WA')"],
            ["• fiscal_year: Format: 'FY####' (e.g., 'FY2026')"],
            ["• dates: Format: 'YYYY-MM-DD' (e.g., '2026-03-15')"],
            ["• phone: Format: '(XXX) XXX-XXXX' or 'XXX-XXX-XXXX'"],
            ["• email: Valid email format"],
            ["• exit_conference_format: 'virtual' or 'in-person'"],
            [""],
            [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        ]

        for row in instructions:
            ws.append(row)

        # Style title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'].font = Font(bold=True)
        ws['A11'].font = Font(bold=True)
        ws['A18'].font = Font(bold=True)

        self._auto_size_columns(ws)

    def _style_header_row(self, ws, row_num: int):
        """Apply styling to header row"""
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_size_columns(self, ws):
        """Auto-size columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def import_from_excel(self, excel_path: str, json_path: str):
        """
        Import Excel template back to JSON configuration.
        """
        print(f"📥 Importing {excel_path} to JSON configuration...")

        # Read all sheets
        excel_file = pd.ExcelFile(excel_path)

        # Read data
        project_info_df = pd.read_excel(excel_file, sheet_name='Project_Info')
        contractors_df = pd.read_excel(excel_file, sheet_name='Contractors')
        fta_pm_df = pd.read_excel(excel_file, sheet_name='FTA_PM')
        regional_officer_df = pd.read_excel(excel_file, sheet_name='Regional_Officer')
        attendees_df = pd.read_excel(excel_file, sheet_name='Attendees')
        subrecipients_df = pd.read_excel(excel_file, sheet_name='Subrecipients')

        # Build projects list
        projects = []

        for _, row in project_info_df.iterrows():
            project_id = int(row['project_id'])

            # Build project object
            project = {
                "project_id": project_id,
                "project_info": {
                    "region_number": int(row['region_number']),
                    "review_type": str(row['review_type']),
                    "recipient_city_state": str(row['recipient_city_state']),
                    "recipient_acronym": str(row.get('recipient_acronym', '')),
                    "recipient_website": str(row.get('recipient_website', '')),
                    "fiscal_year": str(row['fiscal_year']),
                    "site_visit_start_date": str(row['site_visit_start_date']),
                    "site_visit_end_date": str(row['site_visit_end_date']),
                    "site_visit_dates": str(row['site_visit_dates']),
                    "exit_conference_format": str(row['exit_conference_format']),
                    "report_date": str(row.get('report_date', '')) if pd.notna(row.get('report_date')) else ''
                },
                "notes": str(row.get('notes', '')) if pd.notna(row.get('notes')) else ''
            }

            # Add contractor
            contractor_row = contractors_df[contractors_df['project_id'] == project_id].iloc[0]
            project['contractor'] = {
                "lead_reviewer_name": str(contractor_row['lead_reviewer_name']),
                "company_name": str(contractor_row['company_name']),
                "lead_reviewer_phone": str(contractor_row['lead_reviewer_phone']),
                "lead_reviewer_email": str(contractor_row['lead_reviewer_email'])
            }

            # Add FTA PM
            fta_pm_row = fta_pm_df[fta_pm_df['project_id'] == project_id].iloc[0]
            project['fta_program_manager'] = {
                "name": str(fta_pm_row['name']),
                "title": str(fta_pm_row['title']),
                "phone": str(fta_pm_row['phone']),
                "email": str(fta_pm_row['email'])
            }

            # Add Regional Officer
            ro_row = regional_officer_df[regional_officer_df['project_id'] == project_id].iloc[0]
            project['regional_officer'] = {
                "name": str(ro_row['name']),
                "title": str(ro_row['title']),
                "phone": str(ro_row['phone']),
                "email": str(ro_row['email'])
            }

            # Add attendees
            attendees_rows = attendees_df[attendees_df['project_id'] == project_id]
            attendees = []
            for _, att_row in attendees_rows.iterrows():
                if pd.notna(att_row.get('name')) and att_row['name']:
                    attendees.append({
                        "name": str(att_row['name']),
                        "organization": str(att_row['organization']),
                        "title": str(att_row.get('title', '')) if pd.notna(att_row.get('title')) else '',
                        "email": str(att_row.get('email', '')) if pd.notna(att_row.get('email')) else ''
                    })
            project['attendees'] = attendees

            # Add subrecipients
            subrecipients_rows = subrecipients_df[subrecipients_df['project_id'] == project_id]
            subrecipients = {"5307": [], "5310": [], "5311": []}

            for _, sub_row in subrecipients_rows.iterrows():
                if pd.notna(sub_row.get('name')) and sub_row['name']:
                    program_type = str(sub_row['program_type'])
                    subrecipients[program_type].append({
                        "name": str(sub_row['name']),
                        "location": str(sub_row.get('location', '')) if pd.notna(sub_row.get('location')) else ''
                    })

            project['subrecipients'] = subrecipients

            projects.append(project)

        # Build final configuration
        config = {"projects": projects}

        # Save JSON
        with open(json_path, 'w') as f:
            json.dump(config, f, indent=2)

        print(f"✅ JSON configuration saved to: {json_path}")
        print(f"   Projects: {len(projects)}")

    def validate_config(self, json_path: str):
        """Validate JSON configuration against schema"""
        print(f"🔍 Validating {json_path}...")

        try:
            import jsonschema
        except ImportError:
            print("⚠️  jsonschema package not installed. Install with: pip install jsonschema")
            return

        # Load schema
        with open(self.schema_path, 'r') as f:
            schema = json.load(f)

        # Load config
        with open(json_path, 'r') as f:
            config = json.load(f)

        # Validate
        try:
            jsonschema.validate(instance=config, schema=schema)
            print("✅ Configuration is valid!")

            # Print summary
            projects = config.get('projects', [])
            print(f"\nSummary:")
            print(f"  Total projects: {len(projects)}")
            for project in projects:
                project_id = project.get('project_id')
                review_type = project.get('project_info', {}).get('review_type')
                print(f"  - Project {project_id}: {review_type}")

        except jsonschema.exceptions.ValidationError as e:
            print(f"❌ Validation failed: {e.message}")
            print(f"   Path: {' -> '.join(str(p) for p in e.path)}")


def main():
    parser = argparse.ArgumentParser(description="Manage CORTAP project setup configuration")
    subparsers = parser.add_subparsers(dest='command', help='Command to execute')

    # Export command
    export_parser = subparsers.add_parser('export', help='Export JSON to Excel template')
    export_parser.add_argument('--input', default='config/project-setup.json', help='Input JSON file')
    export_parser.add_argument('--output', default='config/project-setup-template.xlsx', help='Output Excel file')

    # Import command
    import_parser = subparsers.add_parser('import', help='Import Excel template to JSON')
    import_parser.add_argument('--input', required=True, help='Input Excel file')
    import_parser.add_argument('--output', default='config/project-setup.json', help='Output JSON file')

    # Validate command
    validate_parser = subparsers.add_parser('validate', help='Validate JSON configuration')
    validate_parser.add_argument('--config', default='config/project-setup.json', help='JSON config file to validate')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    manager = ProjectConfigManager()

    if args.command == 'export':
        manager.export_to_excel(args.input, args.output)
    elif args.command == 'import':
        manager.import_from_excel(args.input, args.output)
    elif args.command == 'validate':
        manager.validate_config(args.config)


if __name__ == '__main__':
    main()
